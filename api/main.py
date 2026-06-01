from fastapi import FastAPI, File, UploadFile, HTTPException, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, Response
from fastapi.encoders import jsonable_encoder
from starlette.concurrency import run_in_threadpool
import asyncio
import tempfile
import os
import json
import uuid
import hashlib
import time
from typing import Optional, Dict, List
from collections import deque
import logging
import base64
from openpyxl import load_workbook
from excel_tool import compare_boms, find_header_row_and_map, read_bom_with_auto_headers, compare_boms_manual
from pdf_tool import extract_bom_from_pdf
from image_tool import extract_bom_from_image, SUPPORTED_EXTS as IMAGE_EXTS
from pydantic import BaseModel
from bom_enhancements import (
    alias_store,
    audit_log,
    apply_alias_folding,
    apply_enhanced_norm_reclassify,
    enhanced_norm,
    extract_revision,
    export_with_template,
)


# ─────────────────────────────────────────────────────────────────────
# In-memory caches & telemetry
# ─────────────────────────────────────────────────────────────────────
# Result cache: SHA-256 of file bytes → (timestamp, result_payload). Keeps
# the last 64 conversions; entries older than 1 hour expire on access.
_RESULT_CACHE: Dict[str, tuple] = {}
_CACHE_MAX = 64
_CACHE_TTL_S = 3600

# Recent error ring buffer for /api/errors (debugging from the UI)
_ERROR_BUFFER: deque = deque(maxlen=50)

# Async PDF conversion jobs. PDF→Excel (esp. the Ollama vision fallback on dense
# drawings) can take minutes — far longer than the edge proxy will hold a single
# request. So the endpoint enqueues a job and the client polls for the result.
# In-memory store; entries expire after 1 hour. Requires a single worker process
# (see Dockerfile CMD) so all polls hit the same store.
_JOBS: Dict[str, dict] = {}
_JOBS_TTL_S = 3600


def _prune_jobs():
    now = time.time()
    for k in [k for k, v in _JOBS.items() if now - v.get("ts", now) > _JOBS_TTL_S]:
        _JOBS.pop(k, None)


def _sha256_of(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _cache_get(key: str):
    item = _RESULT_CACHE.get(key)
    if not item:
        return None
    ts, payload = item
    if time.time() - ts > _CACHE_TTL_S:
        _RESULT_CACHE.pop(key, None)
        return None
    return payload


def _cache_put(key: str, payload):
    if len(_RESULT_CACHE) >= _CACHE_MAX:
        # Evict oldest entry
        oldest = min(_RESULT_CACHE.items(), key=lambda kv: kv[1][0])
        _RESULT_CACHE.pop(oldest[0], None)
    _RESULT_CACHE[key] = (time.time(), payload)


def _record_error(kind: str, message: str, **extra):
    _ERROR_BUFFER.append({
        "ts": time.time(),
        "kind": kind,
        "message": message,
        **extra,
    })


def _build_preview_payload(excel_path: str, output_filename: str, result: Dict, max_preview_rows: int = 200) -> Dict:
    """Read the generated Excel, return JSON with preview rows + base64-encoded workbook."""
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    columns_raw = next(rows_iter, None) or []
    columns = [("" if c is None else str(c)) for c in columns_raw]
    preview_rows = []
    total_rows = 0
    for row in rows_iter:
        total_rows += 1
        if total_rows <= max_preview_rows:
            preview_rows.append([("" if c is None else str(c)) for c in row])
    wb.close()

    with open(excel_path, "rb") as f:
        excel_bytes = f.read()

    return {
        "success": True,
        "filename": output_filename,
        "columns": columns,
        "rows": preview_rows,
        "total_rows": total_rows,
        "preview_truncated": total_rows > max_preview_rows,
        "excel_base64": base64.b64encode(excel_bytes).decode("ascii"),
        "metadata": result.get("metadata", {}),
        "warnings": result.get("warnings", []),
    }


_MIME = {
    "excel": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "word": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
}


def _build_convert_payload(output_path: str, output_filename: str, out_format: str,
                           result: Dict, max_preview_rows: int = 200) -> Dict:
    """Build the converter response from the extractor's own preview rows (works
    for both Excel and Word output — no need to re-read the file for a preview)
    plus the base64-encoded output file and its MIME type."""
    preview = result.get("preview", {}) or {}
    columns = preview.get("columns", [])
    rows = preview.get("rows", [])
    total_rows = preview.get("total_rows", len(rows))

    with open(output_path, "rb") as f:
        file_bytes = f.read()

    return {
        "success": True,
        "filename": output_filename,
        "columns": columns,
        "rows": rows[:max_preview_rows],
        "total_rows": total_rows,
        "preview_truncated": total_rows > max_preview_rows,
        # Kept as `excel_base64` for frontend compatibility — it's the output
        # file bytes regardless of format; `mime` says how to download it.
        "excel_base64": base64.b64encode(file_bytes).decode("ascii"),
        "mime": _MIME.get(out_format, _MIME["excel"]),
        "metadata": result.get("metadata", {}),
        "warnings": result.get("warnings", []),
    }

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Data models for manual column mapping
class ColumnMapping(BaseModel):
    mpn: Optional[str] = None
    qty: Optional[str] = None
    refdes: Optional[str] = None
    description: Optional[str] = None

class ManualComparisonRequest(BaseModel):
    file1_mapping: ColumnMapping
    file2_mapping: ColumnMapping

APP_ENV = os.getenv("APP_ENV", "development").lower()
ALLOWED_ORIGINS = os.getenv("ALLOWED_ORIGINS", "*")

app = FastAPI(
    title="BOM Comparison API",
    description="API for comparing Bill of Materials Excel files",
    version="1.0.0"
)

# Configure CORS
if ALLOWED_ORIGINS == "*":
    app.add_middleware(
        CORSMiddleware,
        allow_origins=["*"],
        allow_credentials=False,
        allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"],
        allow_headers=["*"],
        expose_headers=["*"],
    )
else:
    origins = [o.strip() for o in ALLOWED_ORIGINS.split(",") if o.strip()]
    app.add_middleware(
        CORSMiddleware,
        allow_origins=origins,
        allow_credentials=True,
        allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"],
        allow_headers=["*"]
    )

@app.get("/")
async def root():
    return {"message": "BOM Comparison API is running"}

@app.get("/health")
async def health_check():
    return {"status": "healthy", "service": "bom-comparison-api"}

@app.post("/api/analyze-files")
async def analyze_files(
    file1: UploadFile = File(...),
    file2: UploadFile = File(...)
):
    """
    Analyze Excel files and return available columns for user selection.

    Returns:
        JSON with file columns and suggested mappings
    """
    try:
        # Validate file types
        allowed_extensions = {'.xlsx', '.xls'}
        file1_ext = os.path.splitext(file1.filename)[1].lower()
        file2_ext = os.path.splitext(file2.filename)[1].lower()

        if file1_ext not in allowed_extensions or file2_ext not in allowed_extensions:
            raise HTTPException(
                status_code=400,
                detail="Only .xlsx and .xls files are supported"
            )

        # Create temporary files
        with tempfile.NamedTemporaryFile(delete=False, suffix=file1_ext) as tmp1:
            content1 = await file1.read()
            tmp1.write(content1)
            tmp1.flush()
            tmp1_path = tmp1.name

        with tempfile.NamedTemporaryFile(delete=False, suffix=file2_ext) as tmp2:
            content2 = await file2.read()
            tmp2.write(content2)
            tmp2.flush()
            tmp2_path = tmp2.name

        try:
            # Analyze both files
            logger.info(f"Analyzing columns for {file1.filename} and {file2.filename}")

            # Get columns and auto-detected mappings
            df1, auto_map1 = read_bom_with_auto_headers(tmp1_path)
            df2, auto_map2 = read_bom_with_auto_headers(tmp2_path)

            # Get column lists
            file1_columns = list(df1.columns)
            file2_columns = list(df2.columns)

            # Create suggested mappings based on auto-detection
            file1_suggested = {}
            file2_suggested = {}

            for key, col_name in auto_map1.items():
                file1_suggested[key] = col_name

            for key, col_name in auto_map2.items():
                file2_suggested[key] = col_name

            return JSONResponse(content={
                "file1": {
                    "filename": file1.filename,
                    "columns": file1_columns,
                    "suggested_mapping": file1_suggested
                },
                "file2": {
                    "filename": file2.filename,
                    "columns": file2_columns,
                    "suggested_mapping": file2_suggested
                }
            })

        finally:
            # Clean up temporary files
            try:
                os.unlink(tmp1_path)
                os.unlink(tmp2_path)
            except Exception as e:
                logger.warning(f"Failed to clean up temporary files: {e}")

    except Exception as e:
        import traceback
        logger.error(f"Error during file analysis: {str(e)}")
        logger.error(f"Full traceback: {traceback.format_exc()}")
        raise HTTPException(
            status_code=500,
            detail=f"File analysis failed: {str(e)}"
        )

@app.post("/api/compare-manual")
async def compare_files_manual(
    file1: UploadFile = File(...),
    file2: UploadFile = File(...),
    file1_mpn: str = Form(...),
    file1_qty: str = Form(None),
    file1_refdes: str = Form(None),
    file1_description: str = Form(None),
    file2_mpn: str = Form(...),
    file2_qty: str = Form(None),
    file2_refdes: str = Form(None),
    file2_description: str = Form(None),
    fuzzy_threshold: float = Form(0.88),
    ignore_pairs: Optional[str] = Form(None),
):
    """
    Compare two BOM Excel files using manually selected columns.

    Args:
        file1: First Excel file (original)
        file2: Second Excel file (new version)
        file1_mpn, file1_qty, etc.: Column names for file1
        file2_mpn, file2_qty, etc.: Column names for file2

    Returns:
        JSON with comparison results
    """
    try:
        # Validate file types
        allowed_extensions = {'.xlsx', '.xls'}
        file1_ext = os.path.splitext(file1.filename)[1].lower()
        file2_ext = os.path.splitext(file2.filename)[1].lower()

        if file1_ext not in allowed_extensions or file2_ext not in allowed_extensions:
            raise HTTPException(
                status_code=400,
                detail="Only .xlsx and .xls files are supported"
            )

        # Create temporary files
        with tempfile.NamedTemporaryFile(delete=False, suffix=file1_ext) as tmp1:
            content1 = await file1.read()
            tmp1.write(content1)
            tmp1.flush()
            tmp1_path = tmp1.name

        with tempfile.NamedTemporaryFile(delete=False, suffix=file2_ext) as tmp2:
            content2 = await file2.read()
            tmp2.write(content2)
            tmp2.flush()
            tmp2_path = tmp2.name

        try:
            # Create manual column mappings
            manual_map1 = {}
            manual_map2 = {}

            if file1_mpn: manual_map1['mpn'] = file1_mpn
            if file1_qty: manual_map1['qty'] = file1_qty
            if file1_refdes: manual_map1['refdes'] = file1_refdes
            if file1_description: manual_map1['description'] = file1_description

            if file2_mpn: manual_map2['mpn'] = file2_mpn
            if file2_qty: manual_map2['qty'] = file2_qty
            if file2_refdes: manual_map2['refdes'] = file2_refdes
            if file2_description: manual_map2['description'] = file2_description

            logger.info(f"Manual comparison with mappings:")
            logger.info(f"File1: {manual_map1}")
            logger.info(f"File2: {manual_map2}")

            # Decode rejected pairs + clamp threshold
            rejected: Optional[list] = None
            if ignore_pairs:
                try:
                    parsed = json.loads(ignore_pairs)
                    if isinstance(parsed, list):
                        rejected = [(str(p[0]), str(p[1])) for p in parsed if isinstance(p, list) and len(p) == 2]
                except Exception:
                    rejected = None
            try:
                t = float(fuzzy_threshold)
                t = max(0.5, min(1.0, t))
            except (TypeError, ValueError):
                t = 0.88

            # Perform comparison with manual mappings
            results = compare_boms_manual(
                tmp1_path, tmp2_path, manual_map1, manual_map2,
                fuzzy_threshold=t, ignore_pairs=rejected,
            )

            # Post-process: fold MPN aliases, reclassify formatting-only mods
            results = apply_alias_folding(results, alias_store)
            results = apply_enhanced_norm_reclassify(results)

            # Audit
            try:
                audit_log.record({
                    "kind": "compare-manual",
                    "file1": {"name": file1.filename, "sha256": audit_log.file_hash(tmp1_path)},
                    "file2": {"name": file2.filename, "sha256": audit_log.file_hash(tmp2_path)},
                    "mapping": {"file1": manual_map1, "file2": manual_map2},
                    "summary": results.get("summary_stats", {}),
                    "revision": {
                        "file1": extract_revision(file1.filename),
                        "file2": extract_revision(file2.filename),
                    },
                })
            except Exception as ae:
                logger.warning(f"Audit log write failed: {ae}")

            logger.info("Manual comparison completed successfully")

            return JSONResponse(content=jsonable_encoder(results))

        finally:
            # Clean up temporary files
            try:
                os.unlink(tmp1_path)
                os.unlink(tmp2_path)
            except Exception as e:
                logger.warning(f"Failed to clean up temporary files: {e}")

    except Exception as e:
        import traceback
        logger.error(f"Error during manual comparison: {str(e)}")
        logger.error(f"Full traceback: {traceback.format_exc()}")
        raise HTTPException(
            status_code=500,
            detail=f"Manual comparison failed: {str(e)}"
        )

@app.options("/api/compare")
async def compare_files_options():
    """Handle preflight CORS requests"""
    return {"message": "CORS preflight handled"}

@app.options("/api/analyze-files")
async def analyze_files_options():
    """Handle preflight CORS requests"""
    return {"message": "CORS preflight handled"}

@app.options("/api/compare-manual")
async def compare_manual_options():
    """Handle preflight CORS requests"""
    return {"message": "CORS preflight handled"}

@app.post("/api/compare")
async def compare_files(
    file1: UploadFile = File(...),
    file2: UploadFile = File(...),
    fuzzy_threshold: float = Form(0.88),
    ignore_pairs: Optional[str] = Form(None),
):
    """
    Compare two BOM Excel files and return the differences.
    
    Args:
        file1: First Excel file (original)
        file2: Second Excel file (new version)
    
    Returns:
        JSON with comparison results including:
        - new_parts: Parts only in file2
        - removed_parts: Parts only in file1
        - modified_parts: Parts with BOTH QTY AND REF DES changes (MPN must be the same)
        - unchanged_parts: Parts with identical data
        - unrecognized_parts: Parts that couldn't be categorized
        - summary_stats: Statistics about the comparison
    """
    try:
        # Validate file types
        allowed_extensions = {'.xlsx', '.xls'}
        file1_ext = os.path.splitext(file1.filename)[1].lower()
        file2_ext = os.path.splitext(file2.filename)[1].lower()
        
        if file1_ext not in allowed_extensions or file2_ext not in allowed_extensions:
            raise HTTPException(
                status_code=400,
                detail="Only .xlsx and .xls files are supported"
            )
        
        # Create temporary files with proper flushing
        with tempfile.NamedTemporaryFile(delete=False, suffix=file1_ext) as tmp1:
            content1 = await file1.read()
            tmp1.write(content1)
            tmp1.flush()  # Ensure data is written to disk
            tmp1_path = tmp1.name
        
        with tempfile.NamedTemporaryFile(delete=False, suffix=file2_ext) as tmp2:
            content2 = await file2.read()
            tmp2.write(content2)
            tmp2.flush()  # Ensure data is written to disk
            tmp2_path = tmp2.name
        
        try:
            # Perform comparison
            logger.info(f"Starting comparison of {file1.filename} and {file2.filename}")

            # Decode user-rejected fuzzy pairs JSON: [["MPN_a","MPN_b"], ...]
            rejected: Optional[list] = None
            if ignore_pairs:
                try:
                    parsed = json.loads(ignore_pairs)
                    if isinstance(parsed, list):
                        rejected = [(str(p[0]), str(p[1])) for p in parsed if isinstance(p, list) and len(p) == 2]
                except Exception:
                    rejected = None

            # Clamp threshold
            try:
                t = float(fuzzy_threshold)
                t = max(0.5, min(1.0, t))
            except (TypeError, ValueError):
                t = 0.88

            results = compare_boms(tmp1_path, tmp2_path, fuzzy_threshold=t, ignore_pairs=rejected)

            # Post-process: fold MPN aliases, reclassify formatting-only mods
            results = apply_alias_folding(results, alias_store)
            results = apply_enhanced_norm_reclassify(results)

            # Audit
            try:
                audit_log.record({
                    "kind": "compare-auto",
                    "file1": {"name": file1.filename, "sha256": audit_log.file_hash(tmp1_path)},
                    "file2": {"name": file2.filename, "sha256": audit_log.file_hash(tmp2_path)},
                    "summary": results.get("summary_stats", {}),
                    "revision": {
                        "file1": extract_revision(file1.filename),
                        "file2": extract_revision(file2.filename),
                    },
                })
            except Exception as ae:
                logger.warning(f"Audit log write failed: {ae}")

            # Debug: Log the results before JSON encoding
            logger.info(f"Comparison completed. Removed parts count: {len(results.get('removed_parts', []))}")
            for i, part in enumerate(results.get('removed_parts', [])):
                logger.info(f"  Removed part {i+1}: {part.get('MPN', 'Unknown')}")
            
            logger.info("Comparison completed successfully")
            
            # Ensure JSON-serializable output (handles numpy types etc.)
            payload = jsonable_encoder(results)
            
            # Debug: Log after JSON encoding
            logger.info(f"After JSON encoding. Removed parts count: {len(payload.get('removed_parts', []))}")
            
            return JSONResponse(content=payload)
            
        finally:
            # Clean up temporary files
            try:
                os.unlink(tmp1_path)
                os.unlink(tmp2_path)
            except Exception as e:
                logger.warning(f"Failed to clean up temporary files: {e}")
                
    except Exception as e:
        import traceback
        logger.error(f"Error during comparison: {str(e)}")
        logger.error(f"Full traceback: {traceback.format_exc()}")
        raise HTTPException(
            status_code=500,
            detail=f"Comparison failed: {str(e)}"
        )

def _do_pdf_conversion(content: bytes, filename: str, mode: str, out_format: str) -> Dict:
    """Blocking PDF→file conversion (runs in a threadpool). Writes temp files,
    extracts, builds the preview payload, and cleans up. Raises on failure."""
    with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_pdf:
        tmp_pdf.write(content)
        tmp_pdf.flush()
        pdf_path = tmp_pdf.name

    base, _ = os.path.splitext(filename)
    ext = "docx" if out_format == "word" else "xlsx"
    output_filename = f"{base}.{ext}"
    output_path = os.path.join(tempfile.gettempdir(), f"{uuid.uuid4()}_{output_filename}")
    try:
        result = extract_bom_from_pdf(pdf_path, output_path, mode, out_format)
        return _build_convert_payload(output_path, output_filename, out_format, result)
    finally:
        for p in (pdf_path, output_path):
            try:
                if os.path.exists(p):
                    os.unlink(p)
            except Exception as e:
                logger.warning(f"Failed to clean up temp file {p}: {e}")


async def _run_pdf_job(job_id: str, content: bytes, filename: str,
                       mode: str, out_format: str, cache_key: str):
    """Background task: run the conversion, store the result/error under job_id."""
    try:
        payload = await run_in_threadpool(_do_pdf_conversion, content, filename, mode, out_format)
        _cache_put(cache_key, payload)
        _JOBS[job_id] = {"status": "done", "payload": payload, "ts": time.time()}
        logger.info(f"PDF job {job_id} done: {filename} ({mode}/{out_format})")
    except ValueError as e:
        _record_error("pdf_validation", str(e), filename=filename)
        _JOBS[job_id] = {"status": "error", "detail": str(e), "ts": time.time()}
    except Exception as e:
        import traceback
        logger.error(f"PDF job {job_id} failed: {e}\n{traceback.format_exc()}")
        _record_error("pdf_exception", str(e), filename=filename)
        _JOBS[job_id] = {"status": "error", "detail": f"PDF processing failed: {e}", "ts": time.time()}


@app.post("/api/pdf-to-excel")
async def pdf_to_excel(
    file: UploadFile = File(...),
    mode: str = Form("bom"),
    output_format: str = Form("excel"),
):
    """
    Convert a PDF to Excel/Word — asynchronous.

    mode:   "bom" (normalize to BOM/quote schema + Ollama vision fallback) or
            "normal" (generic any-PDF table, raw).
    output_format: "excel" (.xlsx) or "word" (.docx).

    Returns immediately with {"status": "processing", "job_id": ...}; poll
    GET /api/pdf-to-excel/status/{job_id} for the result. A cached result is
    returned inline as {"status": "done", ...payload}. This keeps long vision
    conversions from exceeding the edge proxy's request timeout.
    """
    if not file.filename.lower().endswith('.pdf'):
        raise HTTPException(status_code=400, detail="File must be PDF format")

    content = await file.read()
    mode = (mode or "bom").lower()
    if mode not in ("bom", "normal"):
        mode = "bom"
    out_format = (output_format or "excel").lower()
    if out_format not in ("excel", "word"):
        out_format = "excel"

    cache_key = f"pdf:{mode}:{out_format}:{_sha256_of(content)}"
    cached = _cache_get(cache_key)
    if cached:
        logger.info(f"Cache hit for PDF {file.filename} ({mode}/{out_format})")
        return JSONResponse(content={"status": "done", **cached})

    _prune_jobs()
    job_id = uuid.uuid4().hex
    _JOBS[job_id] = {"status": "processing", "ts": time.time()}
    logger.info(f"Queued PDF job {job_id}: {file.filename} (mode={mode}, format={out_format})")
    asyncio.create_task(_run_pdf_job(job_id, content, file.filename, mode, out_format, cache_key))
    return JSONResponse(content={"status": "processing", "job_id": job_id}, status_code=202)


@app.get("/api/pdf-to-excel/status/{job_id}")
async def pdf_to_excel_status(job_id: str):
    """Poll a PDF conversion job. Returns {"status": "processing"} |
    {"status": "error", "detail": ...} | {"status": "done", ...payload}."""
    job = _JOBS.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Unknown or expired job id")
    if job["status"] == "processing":
        return JSONResponse(content={"status": "processing"})
    if job["status"] == "error":
        return JSONResponse(content={"status": "error", "detail": job.get("detail", "conversion failed")})
    return JSONResponse(content={"status": "done", **job["payload"]})


@app.options("/api/pdf-to-excel/status/{job_id}")
async def pdf_to_excel_status_options(job_id: str):
    return {"message": "CORS preflight handled"}

@app.options("/api/pdf-to-excel")
async def pdf_to_excel_options():
    """Handle preflight CORS requests"""
    return {"message": "CORS preflight handled"}

@app.post("/api/image-to-excel")
async def image_to_excel(file: UploadFile = File(...)):
    """Extract BOM table from a JPG/PNG screenshot and return a clean Excel file."""
    pdf_path = None
    output_path = None
    try:
        ext = os.path.splitext(file.filename or "")[1].lower()
        if ext not in IMAGE_EXTS:
            raise HTTPException(
                status_code=400,
                detail="File must be .jpg, .jpeg, or .png",
            )

        with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp_img:
            content = await file.read()
            tmp_img.write(content)
            tmp_img.flush()
            pdf_path = tmp_img.name

        cache_key = f"img:{_sha256_of(content)}"
        cached = _cache_get(cache_key)
        if cached:
            logger.info(f"Cache hit for image {file.filename}")
            return JSONResponse(content=cached)

        # Keep the original filename, just swap extension to .xlsx
        output_filename = os.path.splitext(file.filename)[0] + ".xlsx"
        output_path = os.path.join(tempfile.gettempdir(), f"{uuid.uuid4()}_{output_filename}")

        logger.info(f"Processing image: {file.filename}")
        result = await run_in_threadpool(extract_bom_from_image, pdf_path, output_path)
        logger.info(f"Extracted {result['metadata']['rows_extracted']} BOM items from image")

        payload = _build_preview_payload(output_path, output_filename, result)
        _cache_put(cache_key, payload)
        return JSONResponse(content=payload)
    except ValueError as e:
        logger.error(f"Image processing validation error: {e}")
        raise HTTPException(status_code=400, detail=str(e))
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        logger.error(f"Error during image processing: {e}")
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Image processing failed: {e}")
    finally:
        for p in (pdf_path, output_path):
            if p and os.path.exists(p):
                try:
                    os.unlink(p)
                except Exception as cleanup_err:
                    logger.warning(f"Failed to clean up {p}: {cleanup_err}")


@app.options("/api/image-to-excel")
async def image_to_excel_options():
    """Handle preflight CORS requests"""
    return {"message": "CORS preflight handled"}


@app.post("/api/image-to-excel-batch")
async def image_to_excel_batch(files: List[UploadFile] = File(...)):
    """
    Accept N images, extract a BOM from each, and concatenate the rows
    into one Excel file. Headers come from the first image; subsequent
    images contribute data rows aligned by column position.
    """
    if not files:
        raise HTTPException(status_code=400, detail="No files provided")
    import pandas as pd
    from pdf_tool import generate_clean_excel
    paths_to_clean = []
    try:
        all_dfs = []
        warnings_all = []
        for f in files:
            ext = os.path.splitext(f.filename or "")[1].lower()
            if ext not in IMAGE_EXTS:
                raise HTTPException(status_code=400, detail=f"{f.filename}: only .jpg/.jpeg/.png supported")
            with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp_img:
                content = await f.read()
                tmp_img.write(content); tmp_img.flush()
                img_path = tmp_img.name
                paths_to_clean.append(img_path)
            out_xlsx = os.path.join(tempfile.gettempdir(), f"{uuid.uuid4()}_part.xlsx")
            paths_to_clean.append(out_xlsx)
            try:
                result = extract_bom_from_image(img_path, out_xlsx)
                wb = load_workbook(out_xlsx, read_only=True, data_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
                wb.close()
                if not rows:
                    continue
                cols = [("" if c is None else str(c)) for c in rows[0]]
                data = [["" if c is None else str(c) for c in r] for r in rows[1:]]
                df = pd.DataFrame(data, columns=cols)
                df.insert(0, "_source_file", f.filename)
                all_dfs.append(df)
                warnings_all.extend(result.get("warnings", []))
            except Exception as e:
                _record_error("image_batch_part", str(e), filename=f.filename)
                warnings_all.append(f"{f.filename}: {e}")

        if not all_dfs:
            raise HTTPException(status_code=400, detail="No images yielded readable rows")

        merged = pd.concat(all_dfs, ignore_index=True, sort=False).fillna("")
        output_filename = "BOM_merged.xlsx"
        out_path = os.path.join(tempfile.gettempdir(), f"{uuid.uuid4()}_{output_filename}")
        paths_to_clean.append(out_path)
        generate_clean_excel(merged, out_path)
        result = {
            "metadata": {
                "rows_extracted": len(merged),
                "columns_detected": list(merged.columns),
                "files_merged": len(all_dfs),
                "extraction_method": "batch",
                "confidence": 0.7,
            },
            "warnings": warnings_all,
        }
        payload = _build_preview_payload(out_path, output_filename, result)
        return JSONResponse(content=payload)
    finally:
        for p in paths_to_clean:
            try:
                if os.path.exists(p): os.unlink(p)
            except Exception:
                pass


@app.options("/api/image-to-excel-batch")
async def image_to_excel_batch_options():
    return {"message": "CORS preflight handled"}


@app.get("/api/errors")
async def list_recent_errors():
    """Return the last N backend errors. Useful for the UI to surface
    recent failures even when the user has dismissed them."""
    return {"errors": list(_ERROR_BUFFER)}


@app.get("/api/sample-files")
async def list_sample_files():
    """List bundled sample BOMs the UI can offer as a 'Try it' starter."""
    samples_dir = "/app/samples"
    out = []
    if os.path.isdir(samples_dir):
        for name in sorted(os.listdir(samples_dir)):
            full = os.path.join(samples_dir, name)
            if not os.path.isfile(full):
                continue
            ext = os.path.splitext(name)[1].lower()
            kind = "pdf" if ext == ".pdf" else "image" if ext in {".jpg", ".jpeg", ".png"} else "excel"
            out.append({"name": name, "kind": kind, "size": os.path.getsize(full)})
    return {"samples": out}


@app.get("/api/sample-files/{name}")
async def fetch_sample_file(name: str):
    """Stream a bundled sample file by name."""
    from fastapi.responses import FileResponse
    safe = os.path.basename(name)
    full = os.path.join("/app/samples", safe)
    if not os.path.isfile(full):
        raise HTTPException(status_code=404, detail="Sample not found")
    return FileResponse(full, filename=safe)

# ---------------------------------------------------------------------------
# Template-preserving export
# ---------------------------------------------------------------------------

@app.post("/api/export-as-template")
async def export_as_template(
    template_file: UploadFile = File(...),
    data_file: UploadFile = File(...),
    template_mpn: str = Form(...),
    template_qty: Optional[str] = Form(None),
    template_refdes: Optional[str] = Form(None),
    template_description: Optional[str] = Form(None),
    data_mpn: str = Form(...),
    data_qty: Optional[str] = Form(None),
    data_refdes: Optional[str] = Form(None),
    data_description: Optional[str] = Form(None),
    tint: bool = Form(True),
    diff_meta: Optional[str] = Form(None),
):
    """Re-render the data_file's rows inside template_file's styling."""
    allowed = {".xlsx", ".xls"}
    t_ext = os.path.splitext(template_file.filename)[1].lower()
    d_ext = os.path.splitext(data_file.filename)[1].lower()
    if t_ext not in allowed or d_ext not in allowed:
        raise HTTPException(400, "Only .xlsx and .xls files are supported")

    with tempfile.NamedTemporaryFile(delete=False, suffix=t_ext) as tt:
        tt.write(await template_file.read())
        tt.flush()
        template_path = tt.name
    with tempfile.NamedTemporaryFile(delete=False, suffix=d_ext) as td:
        td.write(await data_file.read())
        td.flush()
        data_path = td.name

    out_fd, out_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(out_fd)

    try:
        template_map = {k: v for k, v in {
            "mpn": template_mpn, "qty": template_qty,
            "refdes": template_refdes, "description": template_description,
        }.items() if v}
        data_map = {k: v for k, v in {
            "mpn": data_mpn, "qty": data_qty,
            "refdes": data_refdes, "description": data_description,
        }.items() if v}

        meta_dict = None
        if diff_meta:
            try:
                meta_dict = json.loads(diff_meta)
            except json.JSONDecodeError:
                meta_dict = None

        stats = export_with_template(
            template_path=template_path,
            template_filename=template_file.filename,
            template_mapping=template_map,
            data_path=data_path,
            data_filename=data_file.filename,
            data_mapping=data_map,
            out_path=out_path,
            diff_meta=meta_dict,
            tint=tint,
        )

        with open(out_path, "rb") as fh:
            content = fh.read()

        safe_template_name = os.path.splitext(template_file.filename)[0]
        out_name = f"{safe_template_name}__rewritten.xlsx"

        return Response(
            content=content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{out_name}"',
                "X-Rows-Written": str(stats["rows_written"]),
                "X-Header-Row": str(stats["header_row"]),
            },
        )
    except Exception as e:
        import traceback
        logger.error(f"Template export failed: {e}")
        logger.error(traceback.format_exc())
        raise HTTPException(500, f"Template export failed: {e}")
    finally:
        for p in (template_path, data_path, out_path):
            try:
                if os.path.exists(p):
                    os.unlink(p)
            except Exception:
                pass


@app.options("/api/export-as-template")
async def export_as_template_options():
    return {"message": "CORS preflight handled"}


# ---------------------------------------------------------------------------
# Revision detection
# ---------------------------------------------------------------------------

@app.post("/api/detect-revision")
async def detect_revision(
    file1: UploadFile = File(...),
    file2: UploadFile = File(...),
):
    """Parse 'Rev X' tokens from filenames + first ~20 rows of each workbook."""
    import pandas as pd
    results = {}
    for label, upl in (("file1", file1), ("file2", file2)):
        ext = os.path.splitext(upl.filename)[1].lower()
        with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
            tmp.write(await upl.read())
            tmp.flush()
            path = tmp.name
        try:
            engine = "xlrd" if ext == ".xls" else None
            try:
                if engine:
                    df_raw = pd.read_excel(path, header=None, engine=engine, nrows=20)
                else:
                    df_raw = pd.read_excel(path, header=None, nrows=20)
                rows = df_raw.values.tolist()
            except Exception:
                rows = None
            rev = extract_revision(upl.filename, rows)
            results[label] = {"filename": upl.filename, "revision": rev}
        finally:
            try:
                os.unlink(path)
            except Exception:
                pass
    return results


@app.options("/api/detect-revision")
async def detect_revision_options():
    return {"message": "CORS preflight handled"}


# ---------------------------------------------------------------------------
# MPN alias / cross-reference dictionary
# ---------------------------------------------------------------------------

class AliasIn(BaseModel):
    mpn_a: str
    mpn_b: str
    note: Optional[str] = ""


@app.get("/api/aliases")
async def list_aliases():
    return {"pairs": alias_store.list()}


@app.post("/api/aliases")
async def add_alias(body: AliasIn):
    try:
        pair = alias_store.add(body.mpn_a, body.mpn_b, body.note or "")
        return {"ok": True, "pair": pair}
    except ValueError as e:
        raise HTTPException(400, str(e))


@app.delete("/api/aliases")
async def remove_alias(mpn_a: str, mpn_b: str):
    removed = alias_store.remove(mpn_a, mpn_b)
    return {"ok": removed}


@app.delete("/api/aliases/all")
async def clear_aliases():
    n = alias_store.clear()
    return {"ok": True, "removed": n}


class AliasBatchIn(BaseModel):
    pairs: List[Dict[str, str]]


@app.post("/api/aliases/import")
async def import_aliases(body: AliasBatchIn):
    added = 0
    skipped = 0
    for rec in body.pairs:
        a = (rec.get("mpn_a") or "").strip()
        b = (rec.get("mpn_b") or "").strip()
        if not a or not b or a.upper() == b.upper():
            skipped += 1
            continue
        try:
            alias_store.add(a, b, rec.get("note", ""))
            added += 1
        except Exception:
            skipped += 1
    return {"ok": True, "added": added, "skipped": skipped}


@app.options("/api/aliases")
async def aliases_options():
    return {"message": "CORS preflight handled"}


# ---------------------------------------------------------------------------
# Audit log
# ---------------------------------------------------------------------------

@app.get("/api/audit-log")
async def get_audit_log(limit: int = 100):
    return {"events": audit_log.list(limit=limit)}


@app.delete("/api/audit-log")
async def clear_audit_log():
    audit_log.clear()
    return {"ok": True}


@app.options("/api/audit-log")
async def audit_log_options():
    return {"message": "CORS preflight handled"}


@app.get("/api/test")
async def test_endpoint():
    """Test endpoint to verify API is working"""
    return {
        "message": "API is working",
        "timestamp": "2024-01-01T00:00:00Z",
        "version": "1.0.0"
    }

@app.get("/api/test-comparison")
async def test_comparison():
    """Test endpoint using local files to debug the issue"""
    try:
        file1_path = "test_files/motherboard BOM.xls"
        file2_path = "test_files/MIRO CUBE GEN2 ENT Motherboard REV07A.xls"
        
        # Test direct comparison
        results = compare_boms(file1_path, file2_path)
        
        return {
            "message": "Direct file comparison test",
            "removed_parts_count": len(results.get('removed_parts', [])),
            "removed_parts": [part.get('MPN', 'Unknown') for part in results.get('removed_parts', [])],
            "new_parts_count": len(results.get('new_parts', [])),
            "modified_parts_count": len(results.get('modified_parts', [])),
            "unchanged_parts_count": len(results.get('unchanged_parts', []))
        }
    except Exception as e:
        return {"error": str(e)}

if __name__ == "__main__":
    import uvicorn
    host = "0.0.0.0" if APP_ENV in {"production", "docker"} else "127.0.0.1"
    port = int(os.getenv("PORT", "8000"))
    uvicorn.run(app, host=host, port=port)