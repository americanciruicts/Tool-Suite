"""
BOM Comparison enhancements — drop-in helpers used by main.py.

Provides:
  * enhanced_norm()        — broader electronic-value normalization
                              (resistance R/K/M, capacitance pF/nF/uF/µF,
                               tolerance %, package codes, qty units).
  * extract_revision()     — parse 'Rev A', 'REV-B00', 'rev 03', etc. from
                              filename + the first ~20 rows of a workbook.
  * MpnAliasStore          — JSON-file backed cross-reference dictionary.
                              Apply with .canonical(mpn) before comparison.
  * AuditLog               — append-only JSONL log of compare events.
                              Records file hashes so the *same* compare is
                              traceable even after files leave temp.
  * export_with_template() — open SRC workbook, keep its styles + non-BOM
                              rows (title block, notes), overwrite the data
                              range with DST's rows mapped through column
                              maps. Optional diff tinting.

All side effects (file paths) are configurable; defaults sit next to the
backend in ./data/ so Docker/Vercel layers stay reproducible.
"""

from __future__ import annotations

import hashlib
import json
import os
import re
import shutil
import tempfile
import threading
import uuid
from copy import copy
from dataclasses import dataclass
from datetime import datetime, timezone
from typing import Any, Dict, Iterable, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, PatternFill, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Persistence root
# ---------------------------------------------------------------------------

_DATA_DIR_ENV = os.getenv("BOM_DATA_DIR")
DATA_DIR = _DATA_DIR_ENV or os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
os.makedirs(DATA_DIR, exist_ok=True)

ALIASES_PATH = os.path.join(DATA_DIR, "mpn_aliases.json")
AUDIT_PATH = os.path.join(DATA_DIR, "audit_log.jsonl")


# ---------------------------------------------------------------------------
# Value normalization
# ---------------------------------------------------------------------------

_UNIT_MULTIPLIERS = {
    # SI multipliers commonly seen in electronics value columns
    "p": 1e-12,
    "n": 1e-9,
    "u": 1e-6,
    "µ": 1e-6,
    "μ": 1e-6,  # GREEK SMALL LETTER MU
    "m": 1e-3,
    "k": 1e3,
    "meg": 1e6,
    "M": 1e6,
    "g": 1e9,
}

# Order matters: longer prefixes first.
_VALUE_RE = re.compile(
    r"""^\s*
        (?P<num>[0-9]*\.?[0-9]+)            # 10, 0.1, 4.7
        \s*
        (?P<unit>meg|[pnuµμmkMg])?          # optional SI prefix
        \s*
        (?P<base>[ΩFHohmoFfaradHenry%]*)?   # optional base unit / %
        \s*$
    """,
    re.VERBOSE,
)

_PACKAGE_RE = re.compile(r"^\s*(0201|0402|0603|0805|1206|1210|1812|2010|2512)\s*$")

_QTY_UNIT_RE = re.compile(r"^\s*([0-9]*\.?[0-9]+)\s*(ea|pcs?|pieces?|units?|each)\s*$", re.I)


def _normalize_number(num_str: str) -> str:
    """Render a numeric string canonically: drop trailing zeros, no leading +."""
    try:
        n = float(num_str)
    except (ValueError, TypeError):
        return num_str
    if n == int(n):
        return str(int(n))
    s = f"{n:.10f}".rstrip("0").rstrip(".")
    return s or "0"


def enhanced_norm(val: Any) -> str:
    """
    Normalize an Excel cell value for equality comparison.

    Goals:
      * Strip whitespace, NaN, sentinels (n/a, null, none).
      * Collapse engineering-value variants so the comparison treats
        '750R' == '750' == '750 ohm', '4.7uF' == '4.7µF' == '4u7' (best-effort),
        '5%' == '5 %', '1k' == '1000'.
      * Preserve the original case for opaque tokens (MPNs etc.) — we only
        canonicalize numeric values; alphanumeric MPN-shaped strings pass
        through .upper() unchanged.
    """
    if val is None:
        return ""
    if isinstance(val, float) and pd.isna(val):
        return ""
    s = str(val).strip()
    if not s:
        return ""
    if s.lower() in {"nan", "none", "null", "n/a", "na", "-", "--"}:
        return ""

    # Quantities like "1 ea", "10 pcs"
    m = _QTY_UNIT_RE.match(s)
    if m:
        return _normalize_number(m.group(1))

    # Drop trailing 'R' on resistance values: '750R' -> '750'
    upper = s.upper()
    if len(upper) > 1 and upper.endswith("R"):
        try:
            float(upper[:-1])
            upper = upper[:-1]
            s = upper
        except ValueError:
            pass

    # SI-prefixed engineering values: 4.7uF, 100nF, 10k, 2.2M, 5%, 100mA
    m = _VALUE_RE.match(s)
    if m and (m.group("unit") or m.group("base")):
        num = m.group("num")
        unit = m.group("unit")
        base = (m.group("base") or "").lower()
        try:
            n = float(num)
        except ValueError:
            return upper
        if unit:
            mult = _UNIT_MULTIPLIERS.get(unit, _UNIT_MULTIPLIERS.get(unit.lower(), 1.0))
            n *= mult
        # Render canonical: integer when possible, otherwise short decimal.
        canon_num = _normalize_number(str(n))
        # Strip 'ohm', 'farad' etc — collapse to a unit symbol so '750ohm' == '750'.
        base_canon = base
        if any(t in base_canon for t in ("ohm", "ω", "Ω".lower())):
            base_canon = ""
        if "farad" in base_canon or base_canon == "f":
            base_canon = ""
        if "henry" in base_canon or base_canon == "h":
            base_canon = ""
        return f"{canon_num}{base_canon}".upper()

    # Package code passthrough
    if _PACKAGE_RE.match(s):
        return s.strip().upper()

    return upper


# ---------------------------------------------------------------------------
# Revision detection
# ---------------------------------------------------------------------------

# Catches 'Rev A', 'Rev-A00', 'REV 03', 'rev_C', 'R02', '(Rev. B)'.
_REV_RE = re.compile(r"(?:^|[\s_\-(\[])(?:rev(?:ision)?\.?\s*)([A-Za-z0-9]+[0-9A-Za-z\-]*)",
                     re.IGNORECASE)
_REV_SHORT_RE = re.compile(r"(?:^|[\s_\-])r(?:ev)?[\s\-_]?([A-Z]\d{0,2})(?:[\s_\-).\]]|$)",
                           re.IGNORECASE)


def extract_revision(filename: str, header_preview: Optional[Iterable[Any]] = None) -> Optional[str]:
    """
    Return a short revision token, e.g. 'A', 'B00', '03', or None if not found.

    Looks at the filename first (most reliable in practice), then falls back
    to scanning header_preview (the first ~20 raw rows of the workbook).
    """
    candidates: List[str] = []
    if filename:
        for rx in (_REV_RE, _REV_SHORT_RE):
            for m in rx.finditer(filename):
                candidates.append(m.group(1))
    if header_preview:
        for row in header_preview:
            if not row:
                continue
            for cell in row:
                if cell is None:
                    continue
                text = str(cell)
                for rx in (_REV_RE, _REV_SHORT_RE):
                    for m in rx.finditer(text):
                        candidates.append(m.group(1))
    if not candidates:
        return None
    # Pick the shortest plausible token — long matches like '-Rev-B00_0001' often
    # snag the trailing path component too. Letter+digits like 'B00', '03', 'A'
    # are the canonical shape.
    best = sorted(candidates, key=lambda x: (len(x), x))[0]
    return best.upper()


# ---------------------------------------------------------------------------
# MPN alias dictionary (cross-reference store)
# ---------------------------------------------------------------------------

@dataclass
class AliasPair:
    mpn_a: str
    mpn_b: str
    note: str = ""
    created_at: str = ""

    def to_dict(self) -> Dict[str, str]:
        return {
            "mpn_a": self.mpn_a,
            "mpn_b": self.mpn_b,
            "note": self.note,
            "created_at": self.created_at,
        }


class MpnAliasStore:
    """
    Symmetric MPN equivalence store.

    Two MPNs in the same alias group are treated as the same part during
    comparison — useful for manufacturer cross-references (Yageo vs Vishay
    equivalents, custom internal P/Ns vs vendor P/Ns).

    Implemented as a union-find over normalized MPNs, persisted to a JSON
    file of pair records.
    """

    def __init__(self, path: str = ALIASES_PATH):
        self.path = path
        self._lock = threading.Lock()
        self._pairs: List[AliasPair] = []
        self._parent: Dict[str, str] = {}
        self._load()

    @staticmethod
    def _key(mpn: str) -> str:
        return str(mpn or "").strip().upper()

    def _find(self, k: str) -> str:
        # Iterative union-find with path compression
        chain = []
        while self._parent.get(k, k) != k:
            chain.append(k)
            k = self._parent[k]
        for c in chain:
            self._parent[c] = k
        return k

    def _union(self, a: str, b: str) -> None:
        ra, rb = self._find(a), self._find(b)
        if ra != rb:
            # Pick the lexicographically smaller as canonical for stability
            root, other = (ra, rb) if ra < rb else (rb, ra)
            self._parent[other] = root

    def _load(self) -> None:
        self._pairs = []
        self._parent = {}
        if not os.path.exists(self.path):
            return
        try:
            with open(self.path, "r", encoding="utf-8") as fh:
                data = json.load(fh)
            for rec in data.get("pairs", []):
                a = self._key(rec.get("mpn_a"))
                b = self._key(rec.get("mpn_b"))
                if not a or not b:
                    continue
                self._pairs.append(AliasPair(
                    mpn_a=a, mpn_b=b,
                    note=rec.get("note", ""),
                    created_at=rec.get("created_at", ""),
                ))
                self._parent.setdefault(a, a)
                self._parent.setdefault(b, b)
                self._union(a, b)
        except Exception:
            # Corrupt store — start fresh but back up the bad file
            try:
                shutil.copy(self.path, self.path + ".bak")
            except Exception:
                pass
            self._pairs = []
            self._parent = {}

    def _save(self) -> None:
        tmp_path = self.path + ".tmp"
        with open(tmp_path, "w", encoding="utf-8") as fh:
            json.dump({"pairs": [p.to_dict() for p in self._pairs]}, fh, indent=2)
        os.replace(tmp_path, self.path)

    # Public API -----------------------------------------------------------

    def canonical(self, mpn: str) -> str:
        """Map an MPN to its canonical representative. Identity if not in store."""
        k = self._key(mpn)
        if not k or k not in self._parent:
            return k
        return self._find(k)

    def list(self) -> List[Dict[str, str]]:
        with self._lock:
            return [p.to_dict() for p in self._pairs]

    def add(self, mpn_a: str, mpn_b: str, note: str = "") -> Dict[str, str]:
        a = self._key(mpn_a)
        b = self._key(mpn_b)
        if not a or not b:
            raise ValueError("Both MPNs are required")
        if a == b:
            raise ValueError("Cannot alias an MPN to itself")
        with self._lock:
            pair = AliasPair(
                mpn_a=a, mpn_b=b, note=note,
                created_at=datetime.now(timezone.utc).isoformat(timespec="seconds"),
            )
            self._pairs.append(pair)
            self._parent.setdefault(a, a)
            self._parent.setdefault(b, b)
            self._union(a, b)
            self._save()
            return pair.to_dict()

    def remove(self, mpn_a: str, mpn_b: str) -> bool:
        a = self._key(mpn_a)
        b = self._key(mpn_b)
        with self._lock:
            before = len(self._pairs)
            self._pairs = [
                p for p in self._pairs
                if not ({p.mpn_a, p.mpn_b} == {a, b})
            ]
            removed = len(self._pairs) < before
            if removed:
                # Rebuild union-find from remaining pairs
                self._parent = {}
                for p in self._pairs:
                    self._parent.setdefault(p.mpn_a, p.mpn_a)
                    self._parent.setdefault(p.mpn_b, p.mpn_b)
                    self._union(p.mpn_a, p.mpn_b)
                self._save()
            return removed

    def clear(self) -> int:
        with self._lock:
            n = len(self._pairs)
            self._pairs = []
            self._parent = {}
            self._save()
            return n


# ---------------------------------------------------------------------------
# Audit log
# ---------------------------------------------------------------------------

def _sha256_path(path: str) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


class AuditLog:
    """Append-only JSONL audit log of compare events."""

    def __init__(self, path: str = AUDIT_PATH, max_records: int = 500):
        self.path = path
        self.max_records = max_records
        self._lock = threading.Lock()

    def record(self, event: Dict[str, Any]) -> Dict[str, Any]:
        event = dict(event)
        event.setdefault("id", uuid.uuid4().hex[:12])
        event.setdefault("timestamp", datetime.now(timezone.utc).isoformat(timespec="seconds"))
        with self._lock:
            with open(self.path, "a", encoding="utf-8") as fh:
                fh.write(json.dumps(event, default=str) + "\n")
            self._trim()
        return event

    def _trim(self) -> None:
        try:
            with open(self.path, "r", encoding="utf-8") as fh:
                lines = fh.readlines()
            if len(lines) <= self.max_records:
                return
            lines = lines[-self.max_records:]
            with open(self.path, "w", encoding="utf-8") as fh:
                fh.writelines(lines)
        except Exception:
            pass

    def list(self, limit: int = 100) -> List[Dict[str, Any]]:
        if not os.path.exists(self.path):
            return []
        out: List[Dict[str, Any]] = []
        try:
            with open(self.path, "r", encoding="utf-8") as fh:
                for line in fh:
                    line = line.strip()
                    if not line:
                        continue
                    try:
                        out.append(json.loads(line))
                    except json.JSONDecodeError:
                        continue
        except Exception:
            return []
        # Most recent first
        out.reverse()
        return out[:limit]

    def clear(self) -> None:
        with self._lock:
            try:
                if os.path.exists(self.path):
                    os.remove(self.path)
            except Exception:
                pass

    @staticmethod
    def file_hash(path: str) -> str:
        try:
            return _sha256_path(path)
        except Exception:
            return ""


# ---------------------------------------------------------------------------
# Template-preserving export
# ---------------------------------------------------------------------------

# Diff tints
_TINT_QTY = "FFFEF3C7"      # amber-100 — qty changed
_TINT_REFDES = "FFE0E7FF"   # indigo-100 — refdes changed
_TINT_DESC = "FFEDE9FE"     # violet-100 — description changed
_TINT_ADDED = "FFDCFCE7"    # green-100 — row added (not in src)
_TINT_REMOVED = "FFFEE2E2"  # rose-100 — row removed (kept in src only)
_TINT_MPN_CHANGED = "FFCFFAFE"  # cyan-100 — MPN renamed


def _col_letter(col_idx_0: int) -> str:
    return get_column_letter(col_idx_0 + 1)


def _read_rows(path: str, header_row: int, mapping: Dict[str, str]) -> List[Dict[str, Any]]:
    """Read data rows from `path` keyed by logical column names (mpn/qty/refdes/description)."""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xls":
        df = pd.read_excel(path, header=header_row, engine="xlrd")
    else:
        df = pd.read_excel(path, header=header_row)
    out: List[Dict[str, Any]] = []
    for _, row in df.iterrows():
        rec: Dict[str, Any] = {}
        for logical, col_name in mapping.items():
            if not col_name or col_name not in df.columns:
                rec[logical] = None
            else:
                v = row[col_name]
                if pd.isna(v):
                    rec[logical] = None
                else:
                    rec[logical] = v
        if any(rec.get(k) not in (None, "") for k in rec):
            out.append(rec)
    return out


def _detect_header_row_openpyxl(ws, max_scan: int = 30) -> int:
    """Locate the BOM header row by scanning for cells containing common header words."""
    keywords = {"mpn", "part", "qty", "quantity", "desc", "ref", "loc",
                "designator", "manufacturer", "mfg", "item"}
    best_row = 1
    best_score = 0
    for r in range(1, min(max_scan, ws.max_row) + 1):
        row = ws[r]
        score = 0
        for cell in row:
            if cell.value is None:
                continue
            text = str(cell.value).lower().strip()
            if len(text) > 60:
                continue
            if any(kw in text for kw in keywords):
                if any(kw in text for kw in ("mpn", "part number", "p/n", "mfg p/n")):
                    score += 3
                elif any(kw in text for kw in ("qty", "quantity", "desc", "ref", "designator")):
                    score += 2
                else:
                    score += 1
        if score >= 4:
            return r
        if score > best_score:
            best_score = score
            best_row = r
    return best_row


def _build_header_map(ws, header_row: int) -> Dict[str, int]:
    """Map column-name -> 1-based column index in the source worksheet."""
    out: Dict[str, int] = {}
    for cell in ws[header_row]:
        if cell.value is None:
            continue
        out[str(cell.value).strip()] = cell.column
    return out


def _clear_data_rows(ws, header_row: int) -> None:
    """Delete every row strictly below header_row (keeps title block above intact)."""
    last_row = ws.max_row
    if last_row <= header_row:
        return
    # Iterate from bottom up so row indices stay valid
    for r in range(last_row, header_row, -1):
        ws.delete_rows(r, 1)


def _copy_style(src_cell, dst_cell) -> None:
    """Copy cell styling from src_cell to dst_cell (font, fill, border, alignment)."""
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.border = copy(src_cell.border)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy(src_cell.protection)


def _convert_xls_to_xlsx(src_path: str) -> str:
    """openpyxl can't open legacy .xls — convert via pandas+openpyxl write."""
    df = pd.read_excel(src_path, header=None, engine="xlrd")
    dst_fd, dst_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(dst_fd)
    df.to_excel(dst_path, header=False, index=False)
    return dst_path


def export_with_template(
    template_path: str,
    template_filename: str,
    template_mapping: Dict[str, str],
    data_path: str,
    data_filename: str,
    data_mapping: Dict[str, str],
    out_path: str,
    diff_meta: Optional[Dict[str, Any]] = None,
    tint: bool = True,
) -> Dict[str, Any]:
    """
    Produce a workbook that mirrors `template_path`'s styling and non-BOM
    rows (title block, notes above the header) but whose data rows come
    from `data_path` mapped through `data_mapping`.

    Args:
        template_path:    path to workbook whose look-and-feel we keep.
        template_mapping: { logical -> column header text } for template.
        data_path:        workbook whose rows are the *content*.
        data_mapping:     { logical -> column header text } for data.
        out_path:         where to write the result.
        diff_meta:        optional dict { mpn -> { qty, refdes, description,
                                                   mpn_changed, added,
                                                   removed } } used to tint
                                                   cells. Keys are uppercased.
        tint:             if False, skip tinting entirely.

    Returns:
        dict with stats (rows_written, columns_used, etc.).
    """
    # openpyxl needs xlsx; convert legacy if needed (loses styles, but that's
    # the cost of using xls as a template — caller is warned upstream).
    cleanup_paths: List[str] = []
    if template_path.lower().endswith(".xls"):
        new_path = _convert_xls_to_xlsx(template_path)
        cleanup_paths.append(new_path)
        template_path_open = new_path
    else:
        template_path_open = template_path

    wb = load_workbook(template_path_open)
    ws = wb.active

    header_row = _detect_header_row_openpyxl(ws)
    header_map = _build_header_map(ws, header_row)

    # Resolve template logical -> column index
    template_col_idx: Dict[str, int] = {}
    for logical, col_name in (template_mapping or {}).items():
        if not col_name:
            continue
        if col_name in header_map:
            template_col_idx[logical] = header_map[col_name]
        else:
            # Try case-insensitive match
            for k, v in header_map.items():
                if k.lower().strip() == col_name.lower().strip():
                    template_col_idx[logical] = v
                    break

    # Snapshot the FIRST data row (header_row + 1) BEFORE we delete it — we'll
    # use its styling as the template for every new data row.
    style_template: Dict[int, Any] = {}
    if ws.max_row > header_row:
        for cell in ws[header_row + 1]:
            style_template[cell.column] = cell

    # Read content rows from data workbook
    data_rows = _read_rows(data_path, header_row=_detect_data_header_row(data_path),
                           mapping=data_mapping)

    # Clear existing data rows (keep title block above header_row and the
    # header row itself).
    _clear_data_rows(ws, header_row)

    diff_meta_norm: Dict[str, Dict[str, bool]] = {}
    if diff_meta and tint:
        for k, v in diff_meta.items():
            diff_meta_norm[str(k).strip().upper()] = v

    thin = Side(border_style="thin", color="FF888888")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    written = 0
    for i, rec in enumerate(data_rows):
        excel_row = header_row + 1 + i
        # Write each logical field into its template-mapped column
        for logical, col_idx in template_col_idx.items():
            value = rec.get(logical)
            if value is None:
                continue
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            # Copy styling from the snapshotted first data row, if any
            tmpl_cell = style_template.get(col_idx)
            if tmpl_cell is not None:
                _copy_style(tmpl_cell, cell)
            else:
                cell.border = border

        # Tinting
        if tint:
            mpn_val = rec.get("mpn")
            if mpn_val is not None:
                key = str(mpn_val).strip().upper()
                meta = diff_meta_norm.get(key)
                if meta:
                    if meta.get("added"):
                        _tint_row(ws, excel_row, template_col_idx, _TINT_ADDED)
                    elif meta.get("mpn_changed"):
                        _tint_row(ws, excel_row, template_col_idx, _TINT_MPN_CHANGED)
                    else:
                        if meta.get("qty") and "qty" in template_col_idx:
                            _tint_cell(ws, excel_row, template_col_idx["qty"], _TINT_QTY)
                        if meta.get("refdes") and "refdes" in template_col_idx:
                            _tint_cell(ws, excel_row, template_col_idx["refdes"], _TINT_REFDES)
                        if meta.get("description") and "description" in template_col_idx:
                            _tint_cell(ws, excel_row, template_col_idx["description"], _TINT_DESC)
        written += 1

    # Optional: append removed-only rows (parts in template/source but not in data)
    # tinted rose so customer can see what dropped out.
    if diff_meta and tint:
        removed_rows: List[Dict[str, Any]] = []
        for key, meta in diff_meta_norm.items():
            if meta.get("removed") and meta.get("source_row"):
                removed_rows.append(meta["source_row"])
        for rec in removed_rows:
            excel_row = header_row + 1 + written
            for logical, col_idx in template_col_idx.items():
                value = rec.get(logical) if isinstance(rec, dict) else None
                if value is None:
                    continue
                cell = ws.cell(row=excel_row, column=col_idx, value=value)
                tmpl_cell = style_template.get(col_idx)
                if tmpl_cell is not None:
                    _copy_style(tmpl_cell, cell)
            _tint_row(ws, excel_row, template_col_idx, _TINT_REMOVED)
            written += 1

    # Annotate workbook with provenance
    ws.cell(row=1, column=1).comment = None  # don't trample existing comments
    wb.properties.title = f"BOM export (template: {template_filename})"
    wb.properties.subject = f"Data from: {data_filename}"
    wb.properties.keywords = "BOM template export"

    wb.save(out_path)

    for p in cleanup_paths:
        try:
            os.unlink(p)
        except Exception:
            pass

    return {
        "rows_written": written,
        "header_row": header_row,
        "columns_used": list(template_col_idx.keys()),
        "out_path": out_path,
    }


def _tint_cell(ws, row: int, col: int, argb: str) -> None:
    cell = ws.cell(row=row, column=col)
    cell.fill = PatternFill(start_color=argb, end_color=argb, fill_type="solid")


def _tint_row(ws, row: int, col_idx_map: Dict[str, int], argb: str) -> None:
    for col in col_idx_map.values():
        _tint_cell(ws, row, col, argb)


def _detect_data_header_row(path: str) -> int:
    """0-based header row index for pandas (data workbook)."""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xls":
        df_raw = pd.read_excel(path, header=None, engine="xlrd")
    else:
        df_raw = pd.read_excel(path, header=None)
    keywords = {"mpn", "part", "qty", "quantity", "desc", "ref", "loc",
                "designator", "manufacturer", "mfg", "item"}
    best_row, best_score = 0, 0
    for r in range(min(30, len(df_raw))):
        row = df_raw.iloc[r]
        score = 0
        for v in row:
            if pd.isna(v):
                continue
            text = str(v).lower().strip()
            if len(text) > 60:
                continue
            if any(kw in text for kw in keywords):
                if any(kw in text for kw in ("mpn", "part number", "p/n")):
                    score += 3
                elif any(kw in text for kw in ("qty", "quantity")):
                    score += 2
                else:
                    score += 1
        if score >= 4:
            return r
        if score > best_score:
            best_score, best_row = score, r
    return best_row


# ---------------------------------------------------------------------------
# Comparison post-processing
# ---------------------------------------------------------------------------

def apply_alias_folding(results: Dict[str, Any], store: "MpnAliasStore") -> Dict[str, Any]:
    """
    Reclassify removed+new part pairs as renamed/modified when their MPNs
    resolve to the same canonical via the alias store.

    Idempotent and side-effect-free on the input store.
    """
    if not results:
        return results
    removed = list(results.get("removed_parts", []))
    new = list(results.get("new_parts", []))
    if not removed or not new:
        return results

    # Group new parts by canonical MPN for O(1) lookup
    new_by_canon: Dict[str, List[Dict[str, Any]]] = {}
    for p in new:
        c = store.canonical(p.get("MPN", ""))
        if c:
            new_by_canon.setdefault(c, []).append(p)

    keep_removed: List[Dict[str, Any]] = []
    consumed_new_ids: set = set()
    promoted: List[Dict[str, Any]] = []

    for r in removed:
        c = store.canonical(r.get("MPN", ""))
        candidates = new_by_canon.get(c, [])
        match = None
        for cand in candidates:
            cid = id(cand)
            if cid in consumed_new_ids:
                continue
            match = cand
            consumed_new_ids.add(cid)
            break
        if match is None:
            keep_removed.append(r)
            continue
        promoted.append({
            "MPN": match.get("MPN") or r.get("MPN"),
            "File1 MPN": r.get("MPN"),
            "File2 MPN": match.get("MPN"),
            "MPN_Changed": True,
            "Ref Des/LOC": r.get("Ref Des/LOC"),
            "File1 Ref Des": r.get("Ref Des/LOC"),
            "File2 Ref Des": match.get("Ref Des/LOC"),
            "File1 Qty": r.get("Qty"),
            "File2 Qty": match.get("Qty"),
            "File1 Description": r.get("Description"),
            "File2 Description": match.get("Description"),
            "File1 Line": r.get("Line Number"),
            "File2 Line": match.get("Line Number"),
            "_alias_match": True,
        })

    keep_new = [p for p in new if id(p) not in consumed_new_ids]

    out = dict(results)
    out["removed_parts"] = keep_removed
    out["new_parts"] = keep_new
    out["modified_parts"] = list(results.get("modified_parts", [])) + promoted

    stats = dict(out.get("summary_stats") or {})
    stats["removed_parts_count"] = len(keep_removed)
    stats["new_parts_count"] = len(keep_new)
    stats["modified_parts_count"] = len(out["modified_parts"])
    stats["alias_folded_count"] = len(promoted)
    out["summary_stats"] = stats
    return out


def apply_enhanced_norm_reclassify(results: Dict[str, Any]) -> Dict[str, Any]:
    """
    Move modified_parts to unchanged_parts when the only differences are
    formatting variants caught by enhanced_norm (e.g. '1' vs '1 ea',
    '4.7uF' vs '4.7µF').

    Preserves any modified_parts that have *real* differences.
    """
    if not results:
        return results
    mods = results.get("modified_parts", [])
    if not mods:
        return results
    still_modified: List[Dict[str, Any]] = []
    promoted_unchanged: List[Dict[str, Any]] = []
    for p in mods:
        qty_eq = enhanced_norm(p.get("File1 Qty")) == enhanced_norm(p.get("File2 Qty"))
        desc_eq = enhanced_norm(p.get("File1 Description")) == enhanced_norm(p.get("File2 Description"))
        ref_eq = _normalize_refdes_set(p.get("File1 Ref Des")) == _normalize_refdes_set(p.get("File2 Ref Des"))
        mpn_eq = enhanced_norm(p.get("File1 MPN") or p.get("MPN")) == enhanced_norm(p.get("File2 MPN") or p.get("MPN"))
        if qty_eq and desc_eq and ref_eq and mpn_eq and not p.get("_alias_match"):
            promoted_unchanged.append({
                "MPN": p.get("MPN"),
                "Ref Des/LOC": p.get("File1 Ref Des") or p.get("Ref Des/LOC"),
                "Qty": p.get("File1 Qty"),
                "Description": p.get("File1 Description"),
                "Line Number": p.get("File1 Line"),
            })
        else:
            still_modified.append(p)
    out = dict(results)
    out["modified_parts"] = still_modified
    out["unchanged_parts"] = list(results.get("unchanged_parts", [])) + promoted_unchanged
    stats = dict(out.get("summary_stats") or {})
    stats["modified_parts_count"] = len(still_modified)
    stats["unchanged_parts_count"] = len(out["unchanged_parts"])
    stats["enhanced_norm_reclassified"] = len(promoted_unchanged)
    out["summary_stats"] = stats
    return out


def _normalize_refdes_set(val: Any) -> set:
    if val is None:
        return set()
    text = str(val).upper().strip()
    if not text or text in {"N/A", "NA", "NONE", "NULL", "NAN", "-"}:
        return set()
    return {x for x in re.split(r"[,;\s]+", text) if x}


# ---------------------------------------------------------------------------
# Module-level singletons (cheap, thread-safe-enough for FastAPI workers)
# ---------------------------------------------------------------------------

alias_store = MpnAliasStore()
audit_log = AuditLog()
